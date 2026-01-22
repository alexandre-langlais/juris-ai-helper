import csv
import io

from openpyxl import load_workbook
import xlrd

from app.schemas.models import CSVEntry
from app.logging_config import get_logger

logger = get_logger(__name__)


def parse_excel_xlsx(file_bytes: bytes) -> list[CSVEntry]:
    """
    Parse un fichier Excel XLSX.

    Args:
        file_bytes: Contenu binaire du fichier Excel

    Returns:
        Liste de CSVEntry

    Raises:
        ValueError: Si le format est invalide
    """
    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            raise ValueError("Le fichier Excel ne contient aucune feuille")

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Le fichier Excel est vide")

        return _parse_rows(rows)

    except Exception as e:
        if isinstance(e, ValueError):
            raise
        raise ValueError(f"Erreur lors de la lecture du fichier Excel (xlsx): {e}")


def parse_excel_xls(file_bytes: bytes) -> list[CSVEntry]:
    """
    Parse un fichier Excel XLS (ancien format).

    Args:
        file_bytes: Contenu binaire du fichier Excel

    Returns:
        Liste de CSVEntry

    Raises:
        ValueError: Si le format est invalide
    """
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)

        rows = []
        for row_idx in range(ws.nrows):
            row = [ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)]
            rows.append(tuple(row))

        if not rows:
            raise ValueError("Le fichier Excel est vide")

        return _parse_rows(rows)

    except Exception as e:
        if isinstance(e, ValueError):
            raise
        raise ValueError(f"Erreur lors de la lecture du fichier Excel (xls): {e}")


def _parse_rows(rows: list[tuple]) -> list[CSVEntry]:
    """
    Parse les lignes extraites d'un fichier (CSV ou Excel).

    Args:
        rows: Liste de tuples representant les lignes

    Returns:
        Liste de CSVEntry

    Raises:
        ValueError: Si le format est invalide
    """
    if not rows:
        raise ValueError("Le fichier est vide")

    # Premiere ligne = en-tetes
    headers = [str(h).lower().strip() if h else "" for h in rows[0]]

    # Recherche des colonnes sujet et commentaire
    sujet_idx = None
    commentaire_idx = None

    for idx, header in enumerate(headers):
        if header in ("sujet", "subject", "clause", "topic"):
            sujet_idx = idx
        elif header in ("commentaire", "comment", "annotation", "note"):
            commentaire_idx = idx

    if sujet_idx is None:
        raise ValueError(
            "Colonne 'sujet' non trouvee. "
            "Colonnes acceptees: sujet, subject, clause, topic"
        )
    if commentaire_idx is None:
        raise ValueError(
            "Colonne 'commentaire' non trouvee. "
            "Colonnes acceptees: commentaire, comment, annotation, note"
        )

    # Extraction des entrees
    entries: list[CSVEntry] = []
    for row in rows[1:]:
        if len(row) <= max(sujet_idx, commentaire_idx):
            continue

        sujet = str(row[sujet_idx] or "").strip()
        commentaire = str(row[commentaire_idx] or "").strip()

        if sujet and commentaire:
            entries.append(CSVEntry(sujet=sujet, commentaire=commentaire))

    if not entries:
        raise ValueError("Aucune entree valide trouvee dans le fichier")

    return entries


def parse_spreadsheet(file_bytes: bytes, filename: str) -> list[CSVEntry]:
    """
    Parse un fichier CSV ou Excel selon son extension.

    Args:
        file_bytes: Contenu binaire du fichier
        filename: Nom du fichier (pour determiner le type)

    Returns:
        Liste de CSVEntry

    Raises:
        ValueError: Si le format est invalide
    """
    filename_lower = filename.lower()

    if filename_lower.endswith(".xlsx"):
        logger.info("Parsing fichier Excel XLSX")
        return parse_excel_xlsx(file_bytes)
    elif filename_lower.endswith(".xls"):
        logger.info("Parsing fichier Excel XLS")
        return parse_excel_xls(file_bytes)
    elif filename_lower.endswith(".csv"):
        logger.info("Parsing fichier CSV")
        return parse_csv(file_bytes)
    else:
        raise ValueError(
            "Format de fichier non supporte. "
            "Formats acceptes: .csv, .xlsx, .xls"
        )


def parse_csv(csv_bytes: bytes) -> list[CSVEntry]:
    """
    Parse un fichier CSV contenant les sujets et commentaires.

    Le CSV doit avoir les colonnes: sujet, commentaire
    La première ligne est considérée comme l'en-tête.

    Args:
        csv_bytes: Contenu binaire du fichier CSV

    Returns:
        Liste de CSVEntry

    Raises:
        ValueError: Si le format CSV est invalide
    """
    try:
        # Décodage du CSV (UTF-8 avec fallback Latin-1)
        try:
            content = csv_bytes.decode("utf-8")
        except UnicodeDecodeError:
            content = csv_bytes.decode("latin-1")

        # Lecture du CSV
        reader = csv.DictReader(io.StringIO(content))

        # Vérification des colonnes requises
        if reader.fieldnames is None:
            raise ValueError("Le fichier CSV est vide")

        # Normalisation des noms de colonnes (minuscules, sans espaces)
        normalized_fields = {
            field.lower().strip(): field for field in reader.fieldnames
        }

        # Recherche des colonnes sujet et commentaire
        sujet_col = None
        commentaire_col = None

        for normalized, original in normalized_fields.items():
            if normalized in ("sujet", "subject", "clause", "topic"):
                sujet_col = original
            elif normalized in ("commentaire", "comment", "annotation", "note"):
                commentaire_col = original

        if not sujet_col:
            raise ValueError(
                "Colonne 'sujet' non trouvée. "
                "Colonnes acceptées: sujet, subject, clause, topic"
            )
        if not commentaire_col:
            raise ValueError(
                "Colonne 'commentaire' non trouvée. "
                "Colonnes acceptées: commentaire, comment, annotation, note"
            )

        # Extraction des entrées
        entries: list[CSVEntry] = []
        for row in reader:
            sujet = row.get(sujet_col, "").strip()
            commentaire = row.get(commentaire_col, "").strip()

            if sujet and commentaire:
                entries.append(CSVEntry(sujet=sujet, commentaire=commentaire))

        if not entries:
            raise ValueError("Aucune entrée valide trouvée dans le CSV")

        return entries

    except csv.Error as e:
        raise ValueError(f"Erreur de parsing CSV: {e}")
